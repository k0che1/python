import openpyxl
'''
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Продажи'
sheet.append(['Товар','Продажи','Цена'])
wb.save('D:/study/ранхигс/питон/my_first.xlsx')

wb=openpyxl.load_workbook('D:/study/ранхигс/питон/my_first.xlsx')
ws=wb['Продажи']
ws.append(['Яблоки',100,50])
ws.append(['Груши',75,70])
ws.append(['Сливы',50,90])
wb.save('D:/study/ранхигс/питон/my_first.xlsx')
'''
wb=openpyxl.load_workbook('D:/study/ранхигс/питон/my_first.xlsx')
ws=wb['Продажи']
ws['D1'] = 'Сумма'
total=0
total_price=0
item_count=0
for row in range(2,ws.max_row+1):
    quan=ws[f'B{row}'].value
    price=ws[f'C{row}'].value
    if quan is not None and price is not None:
        amount = quan*price
        ws[f'D{row}'] = amount
        total+=amount
        total_price+=price
        item_count+=1
ws[f'A{ws.max_row+1}']='Итого'
ws[f'D{ws.max_row}']=total
print(total_price/item_count)
wb.save('D:/study/ранхигс/питон/my_first_sales.xlsx')

#поиск максимального
'''
wb=openpyxl.load_workbook('D:/study/ранхигс/питон/my_first_sales.xlsx')
ws=wb['Продажи']'''
column_letter = 'D'  # буква столбца, где ищем максимум
start_row = 2        # если есть заголовок, начинаем со 2‑й строки
max_val = 0
for row in range(start_row, ws.max_row ):
    cell = ws[f'{column_letter}{row}']
    value = cell.value
    num = float(value)
    if num > max_val:
        max_val = num
        max_fr=row

print(f'Товар с максимальной ценой: {ws[f'A{max_fr}'].value}')

#поиск минимального
'''
wb=openpyxl.load_workbook('D:/study/ранхигс/питон/my_first_sales.xlsx')
ws=wb['Продажи']'''
column_letter = 'D'  # буква столбца, где ищем минимум
start_row = 2        # если есть заголовок, начинаем со 2‑й строки
min_val = max_val
for row in range(start_row, ws.max_row ):
    cell = ws[f'{column_letter}{row}']
    value = cell.value
    num = float(value)
    if num < min_val:
        min_val = num
        min_fr=row

print(f'товар с минимальной ценой: {ws[f'A{min_fr}'].value}')
print(f'Количество товаров: {ws.max_row-2}')
